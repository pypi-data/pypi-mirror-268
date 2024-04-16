from libs.enums.ImportType import ImportType
from PyQt5.QtWidgets import QMessageBox
from libs.db import DB
import openpyxl


class ImportFiles:
    def __init__(self):
        self._file_path = None
        self._import_type = None
        self._using_ids = None
        self._db = DB()

    def _error(self, error):
        """
        Выводит ошибку в MessageBox
        :param error: Текст ошибки
        """
        QMessageBox.error(self, "Ошибка", f"{error}", QMessageBox.Ok)

    def import_xls(self):
        """
        Импортироват xls файл в базу данных
        :return: Успешность импорта файла в базу данных
        """
        try:
            if self._file_path is not None:
                sheet = openpyxl.load_workbook(self._file_path).active

                if self._import_type == ImportType.COST_ITEMS:
                    for row in sheet.iter_rows(values_only=True):
                        self._import_to_db(row[0], 'cost_items')
                elif self._import_type == ImportType.PRODUCTS:
                    for row in sheet.iter_rows(values_only=True):
                        self._import_to_db(row[0], 'products')
                elif self._import_type == ImportType.CALCULATION:
                    self._process_calculation_data(sheet)
                elif self._import_type == ImportType.PRODUCTION_PLAN:
                    self._process_production_plan_data(sheet)
                else:
                    return False

            return True
        except Exception as e:
            self._error(e)

    def _process_production_plan_data(self, sheet):
        """
        Обработать данные планы выпуска
        :param sheet: Строчка таблицы
        """
        try:
            for row in sheet.iter_rows(values_only=True):
                if self._using_ids:
                    product_id, quantity = row
                else:
                    product_name, quantity = row
                    product_id = self._db.get_id_by_name(product_name, 'products')[0]

                self._db.cursor.execute(
                    f"INSERT INTO `production_plan`(`product_id`, `quantity`) VALUES ('{product_id}','{quantity}');")
                self._db.connection.commit()
        except Exception as e:
            self._error(e)

    def _process_calculation_data(self, sheet):
        """
        Обработать данные калькуляции
        :param sheet: Строчка таблицы
        """
        try:
            for row in sheet.iter_rows(values_only=True):
                if self._using_ids:
                    product_id, cost_item_id, amount = row
                else:
                    product_name, cost_item_name, amount = row
                    product_id = self._db.get_id_by_name(product_name, 'products')[0]
                    cost_item_id = self._db.get_id_by_name(cost_item_name, 'cost_items')[0]

                self._db.cursor.execute(
                    f"INSERT INTO `calculation`(`product_id`, `cost_item_id`, `amount`) VALUES ('{product_id}','{cost_item_id}','{amount}');")
                self._db.connection.commit()
        except Exception as e:
            self._error(e)

    def set_settings(self, file_path: str, import_type: ImportType = ImportType.COST_ITEMS, using_ids: bool = False):
        """
        Установить настройки
        :param file_path: Путь до файла
        :param import_type: Тип импорта (по умолчанию ImportType.COST_ITEMS)
        :param using_ids: Есть ли в файле id (по умолчанию False)
        """
        self._file_path = file_path
        self._import_type = import_type
        self._using_ids = using_ids

    def _import_to_db(self, name, table):
        """
        Импортировать name в таблицу
        :param name: Значение name
        :param table: Имя таблиыч в базе данных
        :return:
        """
        self._db.cursor.execute(f"INSERT INTO `{table}`(`name`) VALUES ('{name}');")
        self._db.connection.commit()
