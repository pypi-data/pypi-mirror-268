import threading
from zipfile import BadZipFile

import openpyxl
import pandas as pd
import time
from openpyxl.worksheet.worksheet import Worksheet

from pista.core._log_manager import printit
from pista.core.common_service import Commons


class ExcelUtil:
    @staticmethod
    def read_excel_sheet(filepath: str, sheet) -> Worksheet:
        """ Get worksheet obj from excel file """
        try:
            workbook = openpyxl.load_workbook(filepath)
            if sheet is not None and sheet in workbook.sheetnames:
                print('Reading worksheet: ' + sheet)
            else:
                sheet = workbook.sheetnames[0]
                print('Worksheet not provided, considering: ' + sheet)

            worksheet = workbook[sheet]
        except FileNotFoundError as e:
            assert False, 'File not found: ' + filepath
        return worksheet

    @classmethod
    def read_all_excel_rows_for_col(cls, filepath: str, col_header: str, row_header_to_avoid=None) -> str:
        """Returns str with values as csv from all the rows
        """
        thread_id = threading.current_thread().native_id
        col_values_str = ''

        col_hdr_num_dict, row_hdr_num_dict, max_row, max_col = cls._get_all_excel_headers_num_dict(filepath)

        workbook = None
        try:
            workbook = openpyxl.load_workbook(filepath, read_only=False)
            sheet = workbook.sheetnames[0]

            if row_header_to_avoid is not None and str(row_header_to_avoid) in row_hdr_num_dict.keys():
                row_to_avoid = row_hdr_num_dict[str(row_header_to_avoid)]
            else:
                row_to_avoid = None

            for i in range(2, max_row + 1):
                if row_to_avoid is not None and f"{row_to_avoid}" == f"{i}":
                    continue

                col = col_hdr_num_dict[col_header]
                curr_cell_val = workbook[sheet].cell(row=i, column=col).value
                printit(f"xl ({i},{col}) {curr_cell_val}", isToPrint=False)

                if curr_cell_val is not None:
                    if col_values_str == '':
                        col_values_str = curr_cell_val
                    else:
                        col_values_str = col_values_str + ',' + curr_cell_val
        finally:
            if workbook is not None:
                workbook.close()

        # for column in workbook[sheet].iter_cols():
        #     column_name = column[0].value
        #     if column_name == col_header:
        #         for i, cell in enumerate(column):
        #             if i == 0:
        #                 continue
        #             if cell.value is not None and str(cell.value).strip() != '':
        #                 col_values_str = col_values_str + ',' + cell.value

        # print(Commons.build_date_forfilename(), f"{thread_id} reading xl (All,{col_header}): {col_values_str}")
        printit(f"... Thread {thread_id} reading excel data (ALL,{col_header}): {col_values_str}")
        return col_values_str

    @classmethod
    def _get_all_excel_headers_num_dict(cls, filepath: str, sheet: str = None):
        """Gets 2 dicts: column header name with seq num, row header name with seq num
        """
        col_hdr_num_dict = dict()
        row_hdr_num_dict = dict()

        workbook = None
        max_row = None
        max_col = None

        for _ in range(5):
            try:
                workbook = openpyxl.load_workbook(filepath, read_only=True)
                break
            except BadZipFile as e:
                time.sleep(2.0)

        try:
            if workbook is not None:
                sheet = workbook.sheetnames[0] if sheet is None else sheet

                max_col = workbook[sheet].max_column
                max_row = workbook[sheet].max_row

                for i in range(2, max_col + 1):
                    col_header = workbook[sheet].cell(row=1, column=i).value
                    col_hdr_num_dict[col_header] = i

                for i in range(2, max_row + 1):
                    row_header = workbook[sheet].cell(row=i, column=1).value
                    row_hdr_num_dict[str(row_header)] = i
        finally:
            if workbook is not None:
                workbook.close()

        return col_hdr_num_dict, row_hdr_num_dict, max_row, max_col

    @classmethod
    def _get_all_excel_headers_numrange_dict(cls, filepath: str, sheet: str = None):
        """Gets 2 dicts: column header name with seq num, row header name with seq num
        """
        col_hdr_num_dict = dict()
        row_hdr_num_dict = dict()

        workbook = None
        max_row = None

        for _ in range(5):
            try:
                workbook = openpyxl.load_workbook(filepath, read_only=True)
                break
            except BadZipFile as e:
                time.sleep(2.0)

        try:
            if workbook is not None:
                sheet = workbook.sheetnames[0] if sheet is None else sheet

                max_col = workbook[sheet].max_column
                max_row = workbook[sheet].max_row

                i = 2
                while i <= max_col:
                    start = end = i
                    col_header = workbook[sheet].cell(row=1, column=i).value
                    for j in range(start + 1, max_col + 1):
                        extra_col_header = workbook[sheet].cell(row=1, column=j).value
                        if col_header == extra_col_header:
                            end = j
                        else:
                            end = start if end is None else end
                            break
                    i = end + 1
                    col_hdr_num_dict[col_header] = (start, end)

                i = 2
                while i <= max_row:
                    start = end = i
                    row_header = workbook[sheet].cell(row=i, column=1).value
                    for j in range(start + 1, max_row + 1):
                        extra_row_header = workbook[sheet].cell(row=j, column=1).value
                        if row_header == extra_row_header:
                            end = j
                        else:
                            end = start if end is None else end
                            break
                    i = end + 1
                    row_hdr_num_dict[row_header] = (start, end)
        finally:
            if workbook is not None:
                workbook.close()

        return col_hdr_num_dict, row_hdr_num_dict, max_row

    @staticmethod
    def append_to_excel_cell(filepath: str, row_header, col_header: str, cell_value):
        """Appends str to existing cell value
        """
        thread_id = threading.current_thread().native_id

        col_hdr_num_dict, row_hdr_num_dict, max_row, max_col = ExcelUtil._get_all_excel_headers_num_dict(filepath)

        workbook = None

        for _ in range(5):
            try:
                workbook = openpyxl.load_workbook(filepath)
                break
            except BadZipFile as e:
                time.sleep(2.0)

        try:
            # workbook = openpyxl.load_workbook(filepath)
            sheet = workbook.active

            col_hdr_num = col_hdr_num_dict[col_header]
            if str(row_header) in row_hdr_num_dict.keys():
                row_hdr_num = row_hdr_num_dict[str(row_header)]
            else:
                new_row_num = max_row + 1
                # sheet.cell(row=new_row_num, column=1).value = str(row_header)
                sheet.append([str(row_header), '', '', ''])
                row_hdr_num = new_row_num

            curr_cell_val = sheet.cell(row=row_hdr_num, column=col_hdr_num).value
            if curr_cell_val is None or str(curr_cell_val).strip() == '':
                new_cell_val = str(cell_value)
            else:
                new_cell_val = str(curr_cell_val).strip() + ',' + str(cell_value)

            # print(Commons.build_date_forfilename(), f"{thread_id} updating xl ({row_header},{col_header}): {cell_value}")
            printit(f"... Thread {thread_id} updating excel data ({row_header},{col_header}): {cell_value}")
            sheet.cell(row=row_hdr_num, column=col_hdr_num).value = new_cell_val
        finally:
            if workbook is not None:
                workbook.save(filepath)
                workbook.close()

    @staticmethod
    def clear_excel_cells(filepath: str, row_header, col_header: str = None):
        """"""
        thread_id = threading.current_thread().native_id

        col_hdr_num_dict, row_hdr_num_dict, max_row, max_col = ExcelUtil._get_all_excel_headers_num_dict(filepath)

        workbook = None

        for _ in range(5):
            try:
                workbook = openpyxl.load_workbook(filepath)
                break
            except BadZipFile as e:
                time.sleep(2.0)

        try:
            # workbook = openpyxl.load_workbook(filepath)
            sheet = workbook.active

            row_hdr_num = row_hdr_num_dict[str(row_header)] if str(row_header) in row_hdr_num_dict.keys() else None

            if row_hdr_num is not None:
                if col_header is not None:
                    '''Clear only cell for 1 row & 1 column'''
                    # print(Commons.build_date_forfilename(), f"{thread_id} clearing xl ({row_header},{col_header})")
                    printit(f"... Thread {thread_id} clearing excel data ({row_header},{col_header})")

                    col_hdr_num = col_hdr_num_dict[col_header] if str(col_header) in col_hdr_num_dict.keys() else None
                    if col_hdr_num is not None:
                        printit(f"xl ({row_hdr_num},{col_hdr_num})", isToPrint=False)
                        sheet.cell(row=row_hdr_num, column=col_hdr_num).value = ''
                else:
                    '''Clear all cells for 1 row'''
                    # print(Commons.build_date_forfilename(), f"{thread_id} clearing xl ({row_header},)")
                    printit(f"... Thread {thread_id} clearing excel data ({row_header},ALL)")
                    for i in range(2, max_col + 1):
                        printit(f"xl ({row_hdr_num},{i})", isToPrint=False)
                        sheet.cell(row=row_hdr_num, column=i).value = ''
        finally:
            if workbook is not None:
                workbook.save(filepath)
                workbook.close()

    @staticmethod
    def read_excel_col_data(filepath: str, sheet: str, col_header: str) -> dict:
        """Excel file has rows & cols
        rows: attributes (1st row is row header)
        cols: dataset (1st col is col header)
        """
        dictdata = dict()
        if col_header is None or col_header == '':
            assert False, 'Column in variable file is not correct'
        else:
            worksheet = ExcelUtil.read_excel_sheet(filepath, sheet)
            maxcol = worksheet.max_column
            maxrow = worksheet.max_row
            col_index = None
            for c in range(2, maxcol + 1):
                if worksheet.cell(1, c).value == col_header:
                    col_index = c
                    break
            if col_index is None:
                assert False, 'Column in varibale file not found'
            else:
                for r in range(2, maxrow + 1):
                    dictdata[worksheet.cell(r, 1).value] = worksheet.cell(r, col_index).value
        return dictdata

# filepath = 'D:/Practice/AutomationService/resources/data/data_variable.xlsx'
# ws = Excel.read_sheet(filepath, 'DO')
# print(ws)


class ExcelDFUtil:

    @classmethod
    def _get_excel_as_df(cls, file_path: str, sheet_name: str):
        # Load the spreadsheet
        xls = pd.ExcelFile(file_path)
        # Load a sheet into a DataFrame by its name
        df = xls.parse(sheet_name)
        df = df.astype(str)
        df = df.replace({'nan': ''})

        return df

    @classmethod
    def _get_df_attr(cls, df):
        max_row, max_col = df.shape

        row_header_dict = df[df.columns[0]].to_dict()
        final_row_header_dict = {v: k for k, v in row_header_dict.items()}

        final_col_header_dict = {item: index for index, item in enumerate(df.columns.values)}

        return final_row_header_dict, final_col_header_dict, max_row, max_col

    @classmethod
    def read_from_excel(cls, file_path: str, sheet_name: str, row: int, column: int):
        df = cls._get_excel_as_df(file_path, sheet_name)

        # Read data from a particular cell
        data = df.iat[row, column]

        # data = int(float(data)) if Commons.check_number_type(data) == 'float' else data
        data = int(float(data)) if str(data).endswith('.0') else data

        return str(data)

    @classmethod
    def read_from_excel_with_header(cls, file_path, sheet_name, row_header, column_header):
        df = cls._get_excel_as_df(file_path, sheet_name)

        row_header_dict, col_header_dict, max_row, max_col = cls._get_df_attr(df)
        row = row_header_dict[row_header]

        # Read data from a particular cell
        data = df.loc[row, column_header]

        # data = int(float(data)) if Commons.check_number_type(data) == 'float' else data
        data = int(float(data)) if str(data).endswith('.0') else data

        return str(data)

    @classmethod
    def read_all_excel_rows_for_col(cls, filepath: str, col_header: str) -> str:
        """Returns str with values as csv from all the rows
        """
        thread_id = threading.current_thread().native_id
        col_values_str_as_csv = ''

        df = cls._get_excel_as_df(filepath, 'Sheet1')
        row_hdr_num_dict, col_header_dict, max_row, max_col = cls._get_df_attr(df)

        try:
            for i in range(0, max_row):
                curr_cell_val = df.loc[i, col_header]
                printit(f"xl ({i},{col_header}) {curr_cell_val}", isToPrint=False)

                if curr_cell_val and curr_cell_val.lower() != 'nan' and curr_cell_val.strip() != '':
                    if col_values_str_as_csv == '':
                        col_values_str_as_csv = curr_cell_val
                    else:
                        col_values_str_as_csv += ',' + curr_cell_val
        finally:
            pass

        print(Commons.build_date_for_filename(), f"{thread_id} reading xl (All,{col_header}): {col_values_str_as_csv}")
        return col_values_str_as_csv

    @classmethod
    def write_to_excel(cls, file_path: str, sheet_name: str, row: int, column: int, data):
        # data = np.nan if data is None or data == '' else str(data)
        data = int(float(data)) if str(data).endswith('.0') else data

        df = cls._get_excel_as_df(file_path, sheet_name)

        # Write data to a particular cell
        df.iat[row, column] = data
        # Write DataFrame back to Excel file
        df.to_excel(file_path, sheet_name=sheet_name, index=False)

    @classmethod
    def write_to_excel_with_header(cls, file_path: str, sheet_name: str, row_header: str, column_header: str, data):
        # data = np.nan if data is None or data == '' else str(data)
        data = int(float(data)) if str(data).endswith('.0') else data

        df = cls._get_excel_as_df(file_path, sheet_name)

        row_header_dict, col_header_dict, max_row, max_col = cls._get_df_attr(df)
        row = row_header_dict[row_header]

        # Write data to a particular cell
        df.loc[row, column_header] = data
        # Write DataFrame back to Excel file
        df.to_excel(file_path, sheet_name=sheet_name, index=False)

    @classmethod
    def append_to_excel_cell(cls, filepath: str, row_header, col_header: str, cell_value):
        """Appends str to existing cell value
        """
        try:
            # TODO write a new line for THREAD_ID if not found
            df = cls._get_excel_as_df(filepath, 'Sheet1')

            row_header_dict, col_header_dict, max_row, max_col = cls._get_df_attr(df)

            # Write new row_header to excel
            if str(row_header) not in row_header_dict.keys():
                df.loc[len(df.index)] = [row_header, '', '', '', '', '']
                df.to_excel(filepath, 'Sheet1', index=False)

            curr_val = cls.read_from_excel_with_header(filepath, 'Sheet1', row_header, col_header)
            if curr_val and len(curr_val) > 0 and curr_val.strip != '':
                new_val = curr_val + ',' + cell_value
            else:
                new_val = cell_value
            cls.write_to_excel_with_header(filepath, 'Sheet1', row_header, col_header, new_val)
        finally:
            pass

    @classmethod
    def clear_excel_cells(cls, filepath: str, row_header, col_header: str = None):
        """"""
        thread_id = threading.current_thread().native_id

        try:
            df = cls._get_excel_as_df(filepath, 'Sheet1')

            row_header_dict, col_header_dict, max_row, max_col = cls._get_df_attr(df)
            row = row_header_dict[row_header]

            if col_header:
                '''Clear only cell for 1 row & 1 column'''
                print(Commons.build_date_for_filename(), f"{thread_id} clearing xl ({row_header},{col_header})")

                printit(f"xl ({row_header},{col_header})", isToPrint=False)
                cls.write_to_excel_with_header(filepath, 'Sheet1', row_header, col_header, '')
            else:
                '''Clear all cells for 1 row'''
                print(Commons.build_date_for_filename(), f"{thread_id} clearing xl ({row_header},)")

                for c in range(1, max_col):
                    printit(f"xl ({row},{c})", isToPrint=False)
                    cls.write_to_excel(filepath, 'Sheet1', row, c, '')
        finally:
            pass
